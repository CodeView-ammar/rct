from django.shortcuts import render, redirect,get_object_or_404
from .forms import (
   DataImportForm,
   GetAllModel
)  
from django.shortcuts import render
from django.core import serializers
from django.http import HttpResponse
from django.urls import reverse_lazy
from our_core.helper import display_view
from django.http import HttpResponseRedirect,JsonResponse,HttpResponse
from django.utils.translation import ugettext_lazy as _
from django.views.generic import ListView,CreateView,View,UpdateView,DeleteView
from django.urls import reverse
from django.db.models import Q ,Max
from django.http import QueryDict
from django.shortcuts import get_object_or_404

from django_datatables_view.base_datatable_view import BaseDatatableView 

from our_core.our_messages import message
from django.contrib import messages
from ..models import (
    DataImport
     )

from django.db.models import Subquery
from django.db import connection
from django.forms import modelformset_factory
from django.db.models import F
from django.db import transaction, IntegrityError
from dal import autocomplete
from django.db.utils import IntegrityError
from django.utils.html import format_html
from permission.permission.permission import has_screen
from django.views.decorators.csrf import csrf_protect

from django.apps import apps
import openpyxl
import arabic_reshaper
from bidi.algorithm import get_display


import os
from xlwt import Workbook
from django.http import FileResponse, HttpResponseRedirect
import numpy as np
import pandas as pd
from mptt.exceptions import InvalidMove
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.contenttypes.models import ContentType
import re
class ForeignKeyConstraintFailure(IntegrityError):
    def __init__(self, table_name, column_name, referenced_table_name):
        
        column_name=self.column_name
        referenced_table_name=self.referenced_table_name 
        message = f"فشل قيد المفتاح الخارجي: {table_name}.{column_name} يشير إلى {referenced_table_name}"
        super().__init__(message)

def max_number():
    # resulta=DataImport.objects.all().aggregate(Max('number'))
    id_max = DataImport.objects.all().aggregate(Max("number"))["number__max"]
    if id_max == None:
        n =  1
        return str(n)
    else:
        id_max = int(id_max) + 1
        n = id_max
        return str(n)




def read_excel(request):
    file = request.FILES['excel_file']
    global global_data
    global global_file
    df = pd.read_excel(file)
    if 'accounts_type' in df.columns:
        df['accounts_type'] = pd.to_numeric(df['accounts_type'], errors='coerce')
        df['accounts_type'] = df['accounts_type'].astype(pd.Int64Dtype())
        df['accounts_type'] = df['accounts_type'].replace(np.nan, 2)
    df.fillna(0, inplace=True)  # استبدل القيم NaN بقيمة فارغة
    data = df.to_dict(orient='records')
    global_file = file
    global_data = data
    
    return JsonResponse(data, safe=False)


def import_data_from_excel(request):

    global global_file
    global global_data
    document_name=request.POST.get("document_name")
    import_type=request.POST.get("import_type")
    data_={
        "message":"يجب عليك إضافة الملف الإسترداد",
        "status":"0"
    }
    if  global_file:
        file = global_file    
    else:
        data_={"message":"يجب عليك إضافة الملف الإسترداد","status":"0"}
        return data_
    
    if import_type == "1":
        if "accounts" in document_name:
            data_=import_fields_new_data_from_excel_account(global_data,table_names=document_name,request=request)
        else:
            data_=import_fields_new_data_from_excel(global_data,table_names=document_name)
    else:
        data_ =import_fields_from_excel(global_data,table_names=document_name)
       
    return data_

def get_all_models():
    all_models = []
    app_configs = apps.get_app_configs()
    
    for app_config in app_configs:
        models = app_config.get_models()
        all_models.extend(models)
    
    return all_models

def get_app_name(model_name):
    models = get_all_models()
    app_label=[]
    for model in models:
        app_label.append(str(model._meta.app_label))
    unique_list = [x for i, x in enumerate(app_label) if x not in app_label[:i]]
    name_app=None
    for unique_list_ in unique_list:
        try:
            model_instance = apps.get_model(str(unique_list_), model_name)
            name_app=unique_list_
            
            break
        except LookupError:
            pass
    return name_app
    
def export_fields_to_excel(request, file_name, **kwargs):
    type_export = kwargs.get('type_export')
    table_name = kwargs.get('table_names')
    excluded_fields = {"id","created_at","modified_at","created_by","modified_by","english_name","lft","rght","tree_id"	,"level"}  # قائمة بأسماء الحقول المستثناة

    model = apps.get_model(get_app_name(table_name), table_name)
    
    fields = model._meta.get_fields()
    foreign_keys = [field.name for field in fields if field.concrete and field.is_relation and field.name not in excluded_fields]
    
    field_names = [field.name if field.name not in foreign_keys else field.name + "_id" for field in fields if field.concrete and field.name not in excluded_fields]


    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for index, field_name in enumerate(field_names):
        sheet.cell(row=1, column=index+1, value=field_name)

    if type_export == 2 or type_export == 3:
        if type_export == 2:
            data = model.objects.all()[:5]
        else:
            data = model.objects.all()

        for row_index, row_data in enumerate(data, start=2):
            for col_index, field_name in enumerate(field_names, start=1):
                cell_value = getattr(row_data, field_name)
                if cell_value is not None:
                    sheet.cell(row=row_index, column=col_index, value=str(cell_value))
                else:
                    sheet.cell(row=row_index, column=col_index, value=None)

    return workbook




def import_fields_from_excel(data_list, **kwargs):
    table_name = kwargs.get('table_names')
    model = apps.get_model(get_app_name(table_name), table_name)

    try:
        with transaction.atomic():
            for row_data in data_list:
                try:
                    number = row_data.get('number')
                    if number and str(number).isdigit():
                        existing_obj = model.objects.filter(number=number).first()

                        if existing_obj:
                            if 'parent_id' in row_data and row_data['parent_id'] == 0.0:
                                row_data['parent'] = row_data['parent_id'] = None

                            for key, value in row_data.items():
                                setattr(existing_obj, key, value)
                            existing_obj.save()
                        else:
                            obj, created = model.objects.get_or_create(**row_data)
                            if not created:
                                for key, value in row_data.items():
                                    if key == 'parent_id' and str(value) == "0.0":
                                        value = None
                                    setattr(obj, key, value)
                                obj.save()
                    else:
                        data_ = {
                            "message": _("Invalid number value"),
                            "status": "2"
                        }
                        return data_

                except IntegrityError as e:
                    error_message = str(e.args[1])
                    
                    if "1452" in str(e.args[0]) or "Cannot add or update a child row: a foreign key constraint fails" in str(e.args[1]):
                        error_message_ = e.args[0]
                        table_name = re.search(r"`([^`]+)`\.`([^`]+)`", error_message_).group(2)
                        field_name = re.search(r"FOREIGN KEY \(`([^`]+)`\)", error_message_).group(1)

                        data_ = {
                            "message": _("In the {0} table, there is a foreign key constraint on the field {1}. Error in finding a matching value in the associated table.").format(str(table_name), str(field_name)),
                            "status": "2"
                        }
                        return data_
                    elif "(1048, \"Column '" in str(e):
                        field_name = re.search(r"'([^']+)'", str(e)).group(1)

                        data_ = {
                            "message": _("The field {0} is empty and this field is required.").format(str(field_name)),
                            "status": "2"
                        }
                        return data_
                    else:
                        data_ = {
                            "message": _("Error: {0}").format(str(e)),
                            "status": "2"
                        }
                        return data_
   
    except Exception as e:
        print(f"Error: {e}")
        data_ = {
            "message": _("خطاء بحض الحقول ليست صحيحة الرجاء التأكد من الحقول \nاو قم بتحميل نموذج لمعرفة الحقول الإجبارية  \n Error: {0}".format(e)),
            "status": "2"
        }
        return data_

    data_ = {
        "message": _("تمت الإستعادة بنجاح"),
        "status": "1"
    }
    return data_

def import_fields_new_data_from_excel(data_list, **kwargs):
    table_name = kwargs.get('table_names')
    model = apps.get_model(get_app_name(table_name), table_name)
    parent_objects = {}
    obj_re = {}
    try:
        with transaction.atomic():
            for row_data in data_list:
                try:
                    if 'parent_id' in row_data:
                        if row_data['parent_id'] == 0.0:
                            row_data['parent'] = row_data['parent_id'] = None
                        else:
                            row_data['parent'] = None
                            parent_objects[row_data['id']] = row_data['parent_id']
                            row_data['parent_id'] = None
                    if "number" in  row_data:
                        obj, created = model.objects.get_or_create(number=row_data['number'], defaults=row_data)
                        if not created:
                            for key, value in row_data.items():
                                if 'parent_id' in row_data:
                                    if key == 'parent_id':
                                        value = None
                                setattr(obj, key, value)
                            obj.save()
                        else:
                            existing_obj = model.objects.get(number=row_data['number'])
                            for key, value in row_data.items():
                                setattr(existing_obj, key, value)
                            existing_obj.save()
                    else:
                        obj, created = model.objects.get_or_create(**row_data)
                        if not created:
                            for key, value in row_data.items():
                                if key == 'parent_id' and str(value) == "0.0":
                                    value = None
                                setattr(obj, key, value)
                            obj.save()
                        
                except IntegrityError as e:
                    error_message = str(e.args[1])
                    
                        # if "Cannot add or update a child row: a foreign key constraint fails" in error_message:
                        # if error_message.find("Cannot add or update a child row"):
                    if "1452" in str(e.args[0]) or "Cannot add or update a child row: a foreign key constraint fails" in str(e.args[1]):
                        error_message_ = e.args[0]
                        table_name = re.search(r"`([^`]+)`\.`([^`]+)`", error_message_).group(2)
                        field_name = re.search(r"FOREIGN KEY \(`([^`]+)`\)", error_message_).group(1)

                        data_ = {
                            "message": _("في جدول {0} هناك ارتباط (foreign key) بحقل {1}. خطأ في العثور على قيمة مطابقة في الجدول المرتبط بالمعرف.").format(str(table_name), str(field_name)),
                            "status": "2"
                        }
                        return data_
                    if "(1048, \"Column '" in str(e):
                        field_name = re.search(r"'([^']+)'", str(e)).group(1)

                        data_ = {
                            "message": _("\n اسم الحقل {0} فارغ وهذا الحقل إلزامي.").format(str(field_name)),
                            "status": "2"
                        }
                        return data_
                    else:
                        data_ = {
                            "message": _("\n {0}تاكد من البيانات المدخلة.").format(str(e)),
                            "status": "2"
                        }
                        return data_
            if parent_objects:
                for id_, parent_id in parent_objects.items():
                    parent_edit = model.objects.get(id=id_)
                    parent_edit.parent = Accounts.objects.get(id=parent_id)
                    parent_edit.save()

        # transaction.set_rollback(True)
    
    except ValueError as e:
        print(f"Error: {e}")
        data_={
            "message":"يوجد خطاء في البيانات او في الحقول{}".format(e),
            "status":"2"
        }
        return data_


    except Exception as e:
        print(f"Error: {e}")
        data_={
            "message":"{} يوجد خطاء في البيانات او في الحقول".format(str(e)),
            "status":"2"
        }
        return data_

    data_={
            "message":"تمت الإستعادة بنجاح",
            "status":"1"
        }
    return data_


def import_fields_new_data_from_excel_account(data_list, **kwargs):
    table_name = kwargs.get('table_names')
    request = kwargs.get('request')
    
    model = apps.get_model(get_app_name(table_name), table_name)
    parent_objects = {}
    obj_re = {}
    from general_ledger.accounts.forms import AccountsForm
    from general_ledger.models import Accounts,AccountsAndCurrency
    try:
        with transaction.atomic(request.session.get("db_name")):
            for row_data in data_list:
                    
                if "arabic_name" in row_data:
                    arabic_name=row_data['arabic_name']
                else:
                    data_={
                            "message":"يجب ان يحتوي الملف على اسم الحساب arabic_name",
                            "status":"1"
                        }
                    return data_
                
                if "english_name" in row_data:
                    english_name=row_data['english_name']
                else:
                    english_name=''
                if "number" in row_data:
                    number=row_data['number']
                else:
                    data_={
                            "message":"يجب ان يحتوي الملف على number",
                            "status":"1"
                        }
                    return data_
                if "parent" in row_data:
                    if row_data['parent'] == 0:
                        parent =''
                    else:
                        parent_number = Accounts.objects.filter(number=row_data['parent']).values("id")
                        parent = parent_number[0]['id']
                else:
                    data_={
                            "message":"يجب ان يحتوي الملف على parent",
                            "status":"1"
                        }
                    return data_
                if "accounts_type" in row_data:
                    accounts_type=row_data['accounts_type']
                else:
                    data_={
                            "message":"يجب ان يحتوي الملف على accounts_type",
                            "status":"1"
                        }
                    return data_
                if "use_cost_center" in row_data:
                    use_cost_center=row_data['use_cost_center']
                else:
                    use_cost_center=None
                if "account_nature" in row_data:
                    account_nature=row_data['account_nature']
                else:
                    data_={
                            "message":"يجب ان يحتوي الملف على account_nature",
                            "status":"1"
                        }
                    return data_
                if "final_account" in row_data:
                    final_account=row_data['final_account']
                else:
                    data_={
                            "message":"يجب ان يحتوي الملف على final_account",
                            "status":"1"
                        }
                    return data_
                
                if "use_cost_center" in row_data:
                    use_cost_center=row_data['use_cost_center']
                else:
                    use_cost_center="Disabled"
                if "analytical_account" in row_data:
                    analytical_account=row_data['analytical_account']
                else:
                    analytical_account=1
                
                if "currency_id" in row_data:
                    currency_id=row_data['currency_id']
                else:
                    data_={
                            "message":"يجب ان يحتوي الملف على currency_id",
                            "status":"1"
                        }
                    return data_

                
                accounts_=AccountsForm({
                "parent":parent,
                "arabic_name":arabic_name,
                "english_name": english_name,
                "number":str(number),
                "accounts_type": accounts_type,
                "use_cost_center":use_cost_center,
                "account_nature": account_nature,
                "final_account":final_account,
                "analytical_account":analytical_account
                })
                if accounts_.is_valid():
                    accounts = accounts_.save(commit=False)
                    # accounts.created_by_id = request.user.id
                    accounts.save()
                    if accounts.id:
                        AccountsAndCurrency.objects.create(
                                account_id=accounts.id, currency_id=currency_id
                            )   
                    else:
                        raise IntegrityError

                
                else:
                    raise IntegrityError
                    
                # transaction.set_rollback(True)
    except IntegrityError:
        transaction.rollback()
                        
    except ValueError as e:
        print(f"Error: {e}")
        data_={
            "message":"يوجد خطاء في البيانات او في الحقول{}".format(e),
            "status":"2"
        }
        return data_


    except Exception as e:
        print(f"Error: {e}")
        data_={
            "message":"{} يوجد خطاء في البيانات او في الحقول".format(str(e)),
            "status":"2"
        }
        return data_
    

    data_={
            "message":"تمت الإستعادة بنجاح",
            "status":"1"
        }
    return data_


def export_xls(request,type_export,type_file,document_name):
    table_names=document_name
    workbook = export_fields_to_excel(request, "{0}.xlsx".format(document_name), table_names=table_names,type_export=type_export, type_file=type_file)
    if type_file==1:
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{0}.xlsx"'.format(document_name)
        workbook.save(response)
    else:
        df = pd.DataFrame(workbook.active.values)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="{0}.csv"'.format(document_name)
        df.to_csv(response, index=False, header=False)

    return response


################# Start DataImport View ##############
class DataImportView(CreateView):
    """
    Outgoing Order Type View for create new Outgoing Order Type
    """
    def get(self, request, *args, **kwargs):
        
        request.session["top_menu"] = {"Settings": "active"}
        request.session["sup_top_menu"] = {"control_data": "active mm-active"}
        request.session["sub_menu"] = {"data_import": "activeli"}
        account_dic=result={}
        if 'id' in request.GET.keys():
            if request.GET.get('id'):
                try:
                    data=DataImport.objects.filter(pk=request.GET.get('id'))
                  
                    
                    

                    result={'status':1,'data':serializers.serialize('json',data)}
                except:
                   pass
            else:
                
                result={'status':0,'data':''}
            return JsonResponse(result)
        else:
            
            
            branch=request.session.get('branch_id')
            context = {
                'form': DataImportForm(),
                "formmodel":GetAllModel(),    
                'url': reverse('DataImportView'),
                'title_form': _('Import Data'),
                'title_list': _('Import Data'),
                }
            
            return render(request, 'data_import/data_import.html', context)

    def post(self, request, *args, **kwargs):
        form = DataImportForm(request.POST)
        if request.POST.get('id'):
            data=get_object_or_404(DataImport,pk=int(request.POST.get('id')))
            form=DataImportForm(request.POST,instance=data)
        
        if form.is_valid():

            result_import=import_data_from_excel(request)
            obj =[]
            if result_import["status"] == '1':
                obj = form.save()
                msg=message.add_successfully()
                obj.branch_id=request.session.get("branch_id")
                obj.save()
            else:
                msg=message.edit_error()
                result={'status':0,'message':msg,"message2":result_import}
                return JsonResponse(result)         
            if request.POST.get('id'):
                if obj.id:
                    msg=message.update_successfully()
                    result={'status':1,'message':msg,"message2":result_import,"max_number":max_number()}
                else:
                    msg=message.edit_error()
                    result={'status':0,'message':msg,"message2":result_import}
            else:
                if obj.id:
                    msg=message.add_successfully()
                    result={'status':1,'message':msg,"message2":result_import,"max_number":max_number()}
                else:
                    msg=message.add_error(request)
                    result={'status':2,'message':msg,"message2":result_import}
        else:
            result={'status':0,'error':form.errors.as_json()}
        return JsonResponse(result)
          
    def delete(self, request,*args, **kwargs):
        pk=int(QueryDict(request.body).get('id'))
        if pk:
            try:
                data=get_object_or_404(DataImport,pk=pk)
                data.delete()
                msg=message.delete_successfully()
                result={'status':1,'message':msg,"max_number":max_number()}
            except:
                msg=message.delete_error()
                result={'status':0,'message':msg,"max_number":max_number()}
        else:
            msg=message.delete_error()
            result={'status':0,'message':msg,"max_number":max_number()}
        return JsonResponse(result)

################# End DataImport View ##############
################# Satrt DataImport List Json ##############
class DataImportJson(BaseDatatableView):
    """List Customer Groups
    Returns:
        [json] --
    """
    model = DataImport
    columns = [
    'id',  
        'number',
        'document_name',
        'import_type',
        'created_at',
        'action',
        
    ]
    order_columns = [
          'id',  
        'number',
        'document_name',
        'import_type',
        'created_at',
        'action',

    ]
    
    count = 0
    def get_initial_queryset(self):
        
        if has_screen(self.request,'DataImport','view') :#check permission
            return self.model.objects.all()
        else:
            return self.model.objects.none()
       
        
    def render_column(self, row, column):
        """Render Column For Datatable
        """
        if column == "id":
            self.count += 1
            return self.count
        

            
        if column == "action":

            if has_screen(self.request,'DataImport','delete') and has_screen(self.request,'DataImport','edit'):#check permission
                return '<a class="edit_row" data-url="{3}" data-id="{0}" style="DISPLAY: -webkit-inline-box;"  data-toggle="tooltip" title="{1}"><i class="fa fa-edit"></i></a><a class="delete_row" data-url="{3}" data-id="{0}"  style="    DISPLAY: -webkit-inline-box;"     data-toggle="tooltip" title="{2}"><i class="fa fa-trash"></i></a>'.format(row.pk,_("Edit"),_("Delete"),reverse("DataImportView"))
            elif has_screen(self.request,'DataImport','delete'): 
                return '<a class="delete_row" data-url="{2}" data-id="{0}"  style="    DISPLAY: -webkit-inline-box;"     data-toggle="tooltip" title="{1}"><i class="fa fa-trash"></i></a>'.format(row.pk,_("Delete"),reverse("DataImportView"))
            
            elif has_screen(self.request,'DataImport','edit'): 
                return '<a class="edit_row" data-url="{2}" data-id="{0}" style="DISPLAY: -webkit-inline-box;"  data-toggle="tooltip" title="{1}"><i class="fa fa-edit"></i></a>'.format(row.pk,_("Edit"),reverse("DataImportView"))
            else:
                return ''
        else:
            return super(DataImportJson, self).render_column(row, column)

    def filter_queryset(self, qs):
        """To Filter data in table using in search
        """
        sSearch = self.request.GET.get("sSearch", None)
        if sSearch:
            qs = qs.filter(
                   Q(number__icontains=sSearch)
                  | Q(DameIdataimport__icontains=sSearch)                
            )
        return qs



class importGetALLModelAutoComplete(autocomplete.Select2QuerySetView):
    excluded_model = [
        'logentry'
        ,"salesperson"
        ,'calcdepreciation'
        ,'location'
        ,'group'
        ,'permission'
        ,'eventbooktemp'
        ,'eventtype'
        ,'installment'
        ,'installmentdetail'
        ,'reservehall'
        ,'reservehalldetail'
        ,'helpdata'
        ,'helpnotes'
        ,'helpurl'
        ,'contenttype'
        ,'accountingperiod'
        ,'accountingperioddetails'
        ,'activitytype'
        ,'configurationgeneralvariable'
        ,'costcenter'
        ,'costcentertype'
        ,'borkeraccounts'
        ,'closingbranch'
        ,'closingyear'
        ,'generalledgergeneralvariable'
        ,'itemtaxtemplate'
        ,'openingbalances'
        ,'branch'
        ,'company'
        ,'dataexport'
        ,'dataimport'
        ,'notificationvariables'
        ,'activitykey'
        ,'branchsystem'
        ,'grouppartionpermission'
        ,'groupscreens'
        ,'groupsystems'
        ,'grouptabs'
        ,'info_client'
        ,'permissiongroup'
        ,'systems'
        ,'systemtabs'
        ,'tabpartionspermission'
        ,'tabscreens'
        ,'typedevice'
        ,'usersdetiles'
        ,'usersdetilestypedevice'
        ,'usersdetilesuserbranch'
        ,'usersdetilesuserpermissiongroup'
        ,'usersgroup'
        ,'accountingstop'
        ,'save_path_backup'
        ,'discountcodingpurchases'
        ,'purchasedetailsreturns'
        ,'purchaseinvoicelocal'
        ,'purchaseinvoicelocaldetails'
        ,'purchaseinvoicelocalreturns'
        ,'customerreport'
        ,'invoicelayout'
        ,'invoicelayout2'
        ,'labelinvoice'
        ,'layoutbody'
        ,'layoutfooter'
        ,'layoutheader'
        ,'layouttype'
        ,'logoreport'
        ,'settings'
        ,'advancedoptions3'
        ,'burdensalesoperations'
        ,'detailsalesbill'
        ,'detailsalesbillreturn'
        ,'detailssalesbillpayment'
        ,'detailssalesbillreturnpayment'
        ,'invoiceqr'
        ,'quotation'
        ,'quotationdetail'
        ,'salesbill'
        ,'salesbillreturn'
        ,'salesdiscountencoding'
        ,'salesmandata'
        ,'typesalesbill'
        ,'typessalesreturn'
        ,'session'
        ,'additionaldiscountsforpurchaseinvoicelocal'
        ,'additionaldiscountsforpurchaseinvoicelocaldetail'
        ,'generalvariables'
        ,'translation'
        ,'adjustmentsstore'
        ,'adjustmentsstoredetail'
        ,'adjustmentstype'
        ,'incomeorder'
        ,'incomeorderandbills'
        ,'incomeorderdetails'
        ,'incomingordertype'
        ,'itemalternative'
        ,'iteminventory'
        ,'iteminventorydetail'
        ,'itemmovement'
        ,'maingroupitem'
        ,'openingstore'
        ,'outgoingorder'
        ,'outgoingorderdetails'
        ,'outgoingordertype'
        ,'pricetype'
        ,'storegroup'
        ,'subgroupitem'
        ,'suppliersitems'
        ,'transferorder'
        ,'transferorderdetails'
        ,'transferordertype'
        ,'warehousebrokeraccounts'
        ,'warehousevariables']
    
    def get_result_label(self, obj):
        app_label = _(obj.app_label)
        model = _(obj.model)
        return format_html('{} - {}', app_label, model)


    def get_queryset(self):
        qs = ContentType.objects.exclude(model__in=self.excluded_model)
        if self.q:

            qs = qs.filter(
                Q(app_label__icontains=self.q) |
                Q(model__icontains=self.q) 
            )
            
        return qs

    def create_option(self, name, value, label, selected, index, subindex=None, attrs=None):
        option = super().create_option(name, value, label, selected, index, subindex=subindex, attrs=attrs)
        option['value'] = str(value.model)
        return option

    def get_result_value(self, result):
        return str(result.model)

