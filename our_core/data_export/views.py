from django.shortcuts import render, redirect,get_object_or_404
from .forms import (
   DataExportForm,
   GetAllModel
)  
from django.shortcuts import render
from django.core import serializers
from django.http import HttpResponse
from django.urls import reverse_lazy
from our_core.helper import display_view
from django.http import HttpResponseRedirect,JsonResponse,HttpResponse
from django.utils.translation import ugettext_lazy as _
from django.views.generic import ListView,CreateView,View,UpdateView,DeleteView
from django.urls import reverse
from django.db.models import Q ,Max
from django.http import QueryDict
from django.shortcuts import get_object_or_404
from django.contrib.contenttypes.models import ContentType
from django_datatables_view.base_datatable_view import BaseDatatableView 

from our_core.our_messages import message
from django.contrib import messages
from ..models import (
    DataExport
     )

from django.db.models import Subquery
from django.db import connection
from django.forms import modelformset_factory
from django.db.models import F
from django.db import transaction, IntegrityError
from dal import autocomplete

from django.utils.html import format_html
from permission.permission.permission import has_screen
from django.views.decorators.csrf import csrf_protect

from django.apps import apps
import openpyxl
import arabic_reshaper
from bidi.algorithm import get_display


import os
from xlwt import Workbook
from django.http import FileResponse, HttpResponseRedirect




import pandas as pd
from django.http import HttpResponse



def max_number():
    id_max = DataExport.objects.all().aggregate(Max("number"))["number__max"]
    if id_max == None:
        n =  1
        return str(n)
    else:
        id_max = int(id_max) + 1
        n = id_max
        return str(n)



def get_all_models():
    all_models = []
    app_configs = apps.get_app_configs()
    
    for app_config in app_configs:
        models = app_config.get_models()
        all_models.extend(models)
    
    return all_models

def get_app_name(model_name):
    models = get_all_models()
    app_label=[]
    for model in models:
        app_label.append(str(model._meta.app_label))
    unique_list = [x for i, x in enumerate(app_label) if x not in app_label[:i]]
    name_app=None
    for unique_list_ in unique_list:
        try:
            model_instance = apps.get_model(str(unique_list_), model_name)
            name_app=unique_list_
            
            break
        except LookupError:
            pass
    return name_app
    
def export_fields_to_excel(request, file_name, **kwargs):
    table_name = kwargs.get('table_names')
    excluded_fields = {"id","modified_at","created_at","created_by","modified_by","name_en","english_name","lft","rght","tree_id"	,"level"}  # قائمة بأسماء الحقول المستثناة
    model = apps.get_model(get_app_name(table_name), table_name)
    
    checkbox_field_i=request.POST.getlist('checkbox_field')
    
    fields = model._meta.get_fields()
    foreign_keys = [field.name for field in fields if field.concrete and field.is_relation and field.name not in excluded_fields]
    
    field_names = [field.name if field.name not in foreign_keys else field.name + "_id" for field in fields if field.concrete and field.name not in excluded_fields]


    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for index, field_name in enumerate(field_names):
        sheet.cell(row=1, column=index+1, value=field_name)

       
    data = model.objects.all()

    for row_index, row_data in enumerate(data, start=2):
        for col_index, field_name in enumerate(field_names, start=1):
            cell_value = getattr(row_data, field_name)
            if cell_value is not None:
                sheet.cell(row=row_index, column=col_index, value=str(cell_value))
            else:
                sheet.cell(row=row_index, column=col_index, value=None)

    return workbook


def export_xls(request,type_file,document_name):
    global work_done
    
    table_names=document_name
    workbook = export_fields_to_excel(request, "{0}.xlsx".format(document_name), table_names=table_names, type_file=type_file)
    if workbook:
        if type_file==1:
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="{0}.xlsx"'.format(document_name)
            workbook.save(response)
        else:
            df = pd.DataFrame(workbook.active.values)
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="{0}.csv"'.format(document_name)
            df.to_csv(response, index=False, header=False)
        work_done=True
        return response
    else:
        work_done=False
        return False

################# Start DataExport View ##############
class DataExportView(CreateView):
    """
    Outgoing Order Type View for create new Outgoing Order Type
    """
    def get(self, request, *args, **kwargs):
        
        request.session["top_menu"] = {"Settings": "active"}
        request.session["sup_top_menu"] = {"control_data": "active mm-active"}
        request.session["sub_menu"] = {"data_export": "activeli"}
        account_dic=result={}
        if 'id' in request.GET.keys():
            if request.GET.get('id'):
                try:
                    data=DataExport.objects.filter(pk=request.GET.get('id'))
                  
                    
                    

                    result={'status':1,'data':serializers.serialize('json',data)}
                except:
                   pass
            else:
                
                result={'status':0,'data':''}
            return JsonResponse(result)
        else:
            
            
            branch=request.session.get('branch_id')
            context = {
                'form': DataExportForm(),
                "formmodel":GetAllModel(),    
                'url': reverse('DataExportView'),
                'title_form': _('Export Data'),
                'title_list': _('Export Data'),
                }
            
            return render(request, 'data_export/data_export.html', context)

    def post(self, request, *args, **kwargs):
        form = DataExportForm(request.POST)
        if request.POST.get('id'):
            data=get_object_or_404(DataExport,pk=int(request.POST.get('id')))
            form=DataExportForm(request.POST,instance=data)
        if form.is_valid():
            
            checkbox_values = request.POST.getlist('checkbox_field')
            
            if not checkbox_values:
                
                result={'status':2,'message':"يجب إختيار الحقول","message2":"error"}
                return 

            if export_xls(request,request.POST.get('type_file'),request.POST.get('document_name')):
                obj = form.save()
                obj.branch_id=request.session.get("branch_id")
                obj.save()
                if request.POST.get('id'):
                    if obj.id:
                        msg=message.update_successfully()
                        result={'status':1,'message':msg,"message2":"تمت التصدير بنجاح","max_number":max_number()}
                    else:
                        msg=message.edit_error()
                        result={'status':2,'message':msg,"message2":"error"}
                else:
                    if obj.id:
                        msg=message.add_successfully()
                        result={'status':1,'message':msg,"message2":"تمت التصدير بنجاح","max_number":max_number()}
                    else:
                        msg=message.add_error(request)
                        result={'status':2,'message':msg,"message2":""}
            else:
                
                result={'status':0,'message':"اسم الشاشة يجب ان يكون صحيح وان يكون بالغة الانجليزية"}
                
        else:
            result={'status':0,'error':form.errors.as_json()}
        return JsonResponse(result)
          
    def delete(self, request,*args, **kwargs):
        pk=int(QueryDict(request.body).get('id'))
        if pk:
            try:
                data=get_object_or_404(DataExport,pk=pk)
                data.delete()
                msg=message.delete_successfully()
                result={'status':1,'message':msg,"max_number":max_number()}
            except:
                msg=message.delete_error()
                result={'status':0,'message':msg,"max_number":max_number()}
        else:
            msg=message.delete_error()
            result={'status':0,'message':msg,"max_number":max_number()}
        return JsonResponse(result)

from django.apps import apps

def get_required_fields(fields):
    
    required_fields = []
    for field in fields:
        if field.concrete and not field.blank and field.name != 'id' and not field.is_relation:
            required_fields.append(field.verbose_name)

    return required_fields

from django.core import serializers

def get_filde_document_name(request):
    model = apps.get_model(get_app_name(request.POST.get('document_name')), request.POST.get('document_name'))
    fields = model._meta.get_fields()
    required_fields = get_required_fields(fields)
    excluded_fields = {"id", "ID", "modified_at", "created_at", "created_by", "modified_by", "lft", "rght", "tree_id", "level"}  # قائمة بأسماء الحقول المستثناة
    foreign_keys = [field.verbose_name for field in fields if field.concrete and field.is_relation and field.verbose_name not in excluded_fields]

    field_names = []
    for field in fields:
        if field.concrete and field.verbose_name not in excluded_fields:
            field_name = field.verbose_name if field.name not in foreign_keys else field.name + "_id"
            if field_name in required_fields:  # تحقق مما إذا كان الحقل إجباريًا
                field_name += " (إجباري)"  # إضافة كلمة "إجباري" بجانب اسم الحقل
            field_names.append({ field.name:field_name})

    result = {'status': 1, 'data': field_names}



    return JsonResponse(result)

################# End DataExport View ##############
################# Satrt DataExport List Json ##############
class DataExportJson(BaseDatatableView):
    """List Customer Groups
    Returns:
        [json] --
    """
    model = DataExport
    columns = [
    'id',  
        'number',
        'document_name',
        'type_file',
        'created_at',
        'action',
        
    ]
    order_columns = [
          'id',  
        'number',
        'document_name',
        'type_file',
        'created_at',
        'action',

    ]
    
    count = 0
    def get_initial_queryset(self):
        
        if has_screen(self.request,'DataExport','view') :#check permission
            return self.model.objects.all()
        else:
            return self.model.objects.none()
       
        
    def render_column(self, row, column):
        """Render Column For Datatable
        """
        if column == "id":
            self.count += 1
            return self.count
        

            
        if column == "action":

            if has_screen(self.request,'DataExport','delete'): 
                return '<a class="delete_row" data-url="{2}" data-id="{0}"  style="    DISPLAY: -webkit-inline-box;"     data-toggle="tooltip" title="{1}"><i class="fa fa-trash"></i></a>'.format(row.pk,_("Delete"),reverse("DataExportView"))
            else:
                return ''
        else:
            return super(DataExportJson, self).render_column(row, column)

    def filter_queryset(self, qs):
        """To Filter data in table using in search
        """
        sSearch = self.request.GET.get("sSearch", None)
        if sSearch:
            qs = qs.filter(
                   Q(number__icontains=sSearch)
                  | Q(DameIDataExport__icontains=sSearch)                
            )
        return qs


from django.utils.translation import ugettext_lazy as _
from django.utils.translation import ugettext
class GetALLModelAutoComplete(autocomplete.Select2QuerySetView):
    excluded_model = [
        'logentry'
        
        ,'calcdepreciation'
        ,'location'
        ,'group'
        ,'permission'
        ,'eventbooktemp'
        ,'eventtype'
        ,'installment'
        ,'installmentdetail'
        ,'reservehall'
        ,'reservehalldetail'
        ,'helpdata'
        ,'helpnotes'
        ,'helpurl'
        ,'contenttype'
        ,'accountingperiod'
        ,'accountingperioddetails'
        ,'activitytype'
        ,'configurationgeneralvariable'
        ,'costcenter'
        ,'costcentertype'
        ,'borkeraccounts'
        ,'closingbranch'
        ,'closingyear'
        ,'generalledgergeneralvariable'
        ,'itemtaxtemplate'
        ,'openingbalances'
        ,'branch'
        ,'company'
        ,'dataexport'
        ,'dataimport'
        ,'notificationvariables'
        ,'activitykey'
        ,'branchsystem'
        ,'grouppartionpermission'
        ,'groupscreens'
        ,'groupsystems'
        ,'grouptabs'
        ,'info_client'
        ,'permissiongroup'
        ,'systems'
        ,'systemtabs'
        ,'tabpartionspermission'
        ,'tabscreens'
        ,'typedevice'
        ,'usersdetiles'
        ,'usersdetilestypedevice'
        ,'usersdetilesuserbranch'
        ,'usersdetilesuserpermissiongroup'
        ,'usersgroup'
        ,'accountingstop'
        ,'save_path_backup'
        ,'discountcodingpurchases'
        ,'purchasedetailsreturns'
        ,'purchaseinvoicelocal'
        ,'purchaseinvoicelocaldetails'
        ,'purchaseinvoicelocalreturns'
        ,'customerreport'
        ,'invoicelayout'
        ,'invoicelayout2'
        ,'labelinvoice'
        ,'layoutbody'
        ,'layoutfooter'
        ,'layoutheader'
        ,'layouttype'
        ,'logoreport'
        ,'settings'
        ,'advancedoptions3'
        ,'burdensalesoperations'
        ,'detailsalesbill'
        ,'detailsalesbillreturn'
        ,'detailssalesbillpayment'
        ,'detailssalesbillreturnpayment'
        ,'invoiceqr'
        ,'quotation'
        ,'quotationdetail'
        ,'salesbill'
        ,'salesbillreturn'
        ,'salesdiscountencoding'
        ,'salesmandata'
        ,'typesalesbill'
        ,'typessalesreturn'
        ,'session'
        ,'additionaldiscountsforpurchaseinvoicelocal'
        ,'additionaldiscountsforpurchaseinvoicelocaldetail'
        ,'generalvariables'
        ,'translation'
        ,'adjustmentsstore'
        ,'adjustmentsstoredetail'
        ,'adjustmentstype'
        ,'incomeorder'
        ,'incomeorderandbills'
        ,'incomeorderdetails'
        ,'incomingordertype'
        ,'itemalternative'
        ,'iteminventory'
        ,'iteminventorydetail'
        ,'itemmovement'
        ,'maingroupitem'
        ,'openingstore'
        ,'outgoingorder'
        ,'outgoingorderdetails'
        ,'outgoingordertype'
        ,'pricetype'
        ,'storegroup'
        ,'subgroupitem'
        ,'suppliersitems'
        ,'transferorder'
        ,'transferorderdetails'
        ,'transferordertype'
        ,'warehousebrokeraccounts'
        ,'warehousevariables']
    
    def get_result_label(self, obj):
        app_label = _(obj.app_label)
        model = _(obj.model)
        return format_html('{} - {}', app_label, model)


    def get_queryset(self):
        qs = ContentType.objects.all().exclude(model__in=self.excluded_model)
        if self.q:
            translated_q =ugettext(self.q)

            qs = qs.filter(
                Q(app_label__icontains=self.q) |
                Q(model__icontains=self.q) |
                Q(model__icontains=translated_q)
            )
            
        return qs

    def create_option(self, name, value, label, selected, index, subindex=None, attrs=None):
        option = super().create_option(name, value, label, selected, index, subindex=subindex, attrs=attrs)
        option['value'] = str(value.model)
        return option

    def get_result_value(self, result):
        return str(result.model)
