from django.shortcuts import render

# Create your views here.
from django.http import HttpResponse
from openpyxl import Workbook


from django.http import HttpResponse
from openpyxl import load_workbook
import io
import datetime
from .models import SupervisoryVisits,InstructorEvaluation
def download_excel(request,id_Supervisory):
    
    if id_Supervisory:
        get_data=SupervisoryVisits.objects.filter(id=id_Supervisory).first()
        if get_data:

            template_file = 'static/template.xlsx'
            workbook = load_workbook(template_file)
            worksheet = workbook.active
            # الفصل التدريبي
            worksheet['A4'] = f"الفصل التدريبي {get_data.year}"
            
            ojb=InstructorEvaluation.objects.filter(supervisoryvisits_id=id_Supervisory).values(
                "planning_1",
                "planning_2",
                "competencies_1",
                "competencies_2",
                "competencies_3",
                "competencies_4",
                "competencies_5",
                "competencies_6",
                "competencies_7",
                "competencies_8",
                "competencies_9",
                "competencies_10",
                "competencies_11",
                "competencies_12",
                "competencies_13",
                "competencies_14",
                "competencies_15",
                "competencies_16",
                "competencies_17",
                "competencies_18",
                "competencies_19",
                "competencies_20",
                "competencies_21",
                "competencies_22",
                "competencies_23",
                "competencies_24",
                "competencies_25",
                "competencies_26",
                "competencies_27",
                "development_1",
                "development_2",
                "development_3",
                "development_4",
                "development_5",
                "development_6",
                "productivity_1",
                "productivity_2",
                "productivity_3",
                "productivity_4",
                "productivity_5",
                "productivity_6",
                "productivity_7",
                "productivity_8",
                "productivity_9",
                "productivity_10",
                "productivity_11",
                "productivity_12",
                "productivity_13",
                "productivity_14",
                "productivity_15",
                "productivity_16",
                "productivity_17",
                "productivity_18",
                "productivity_19",
                "productivity_20",
                "activities_1",
                "activities_2",
                "activities_3",
                "personal_1",
                "personal_2",
                "personal_3",
                "personal_4",
                "personal_5",
                "personal_6",
                "personal_7",
                "personal_8",
                "personal_9",
                "personal_10",
                "personal_11",
                "personal_12",
                "personal_13",
                "personal_14",
                "personal_15",
                "personal_16",
                "personal_17",
                "personal_18",
                "personal_19",
                "personal_20",
                "personal_21",
                "personal_22",
                "personal_23",
            )[0]
            # تعديل خلايا الملف
            worksheet['A5'] = 'مجلس التدريب التقني والمهني اليمن'
            worksheet['A9'] =get_data.trainer.fullname
            worksheet['E9'] =get_data.trainer.id
            worksheet['K9'] =get_data.trainer.date_start
            worksheet['K6'] =get_data.trainer.Cuntry_study.name_cuntry
            worksheet['A12'] =get_data.trainer.qualification
            
             
            worksheet['G18']=ojb["planning_1"]
            worksheet['G19']=ojb["planning_2"]
            worksheet['G21']=ojb["competencies_1"]
            worksheet['G22']=ojb["competencies_2"]
            worksheet['G23']=ojb["competencies_3"]
            worksheet['G24']=ojb["competencies_4"]
            worksheet['G25']=ojb["competencies_5"]
            worksheet['G26']=ojb["competencies_6"]
            worksheet['G27']=ojb["competencies_7"]
            worksheet['G28']=ojb["competencies_8"]
            worksheet['G29']=ojb["competencies_9"]
            worksheet['G30']=ojb["competencies_10"]
            worksheet['G31']=ojb["competencies_11"]
            worksheet['G32']=ojb["competencies_12"]
            worksheet['G33']=ojb["competencies_13"]
            worksheet['G34']=ojb["competencies_14"]
            worksheet['G35']=ojb["competencies_15"]
            worksheet['G36']=ojb["competencies_16"]
            worksheet['G37']=ojb["competencies_17"]
            worksheet['G38']=ojb["competencies_18"]
            worksheet['G39']=ojb["competencies_19"]
            worksheet['G40']=ojb["competencies_20"]
            worksheet['G41']=ojb["competencies_21"]
            worksheet['G42']=ojb["competencies_22"]
            worksheet['G43']=ojb["competencies_23"]
            worksheet['G44']=ojb["competencies_24"]
            worksheet['G45']=ojb["competencies_25"]
            worksheet['G46']=ojb["competencies_26"]
            worksheet['G47']=ojb["competencies_27"]
            worksheet['G49']=ojb["development_1"]
            worksheet['G50']=ojb["development_2"]
            worksheet['G51']=ojb["development_3"]
            worksheet['G52']=ojb["development_4"]
            worksheet['G53']=ojb["development_5"]
            worksheet['G54']=ojb["development_6"]
            worksheet['G56']=ojb["productivity_1"]
            worksheet['G57']=ojb["productivity_2"]
            worksheet['G58']=ojb["productivity_3"]
            worksheet['G59']=ojb["productivity_4"]
            worksheet['G60']=ojb["productivity_5"]
            worksheet['G61']=ojb["productivity_6"]
            worksheet['G62']=ojb["productivity_7"]
            worksheet['G63']=ojb["productivity_8"]
            worksheet['G64']=ojb["productivity_9"]
            worksheet['G65']=ojb["productivity_10"]
            worksheet['G66']=ojb["productivity_11"]
            worksheet['G67']=ojb["productivity_12"]
            worksheet['G68']=ojb["productivity_13"]
            worksheet['G69']=ojb["productivity_14"]
            worksheet['G70']=ojb["productivity_15"]
            worksheet['G71']=ojb["productivity_16"]
            worksheet['G72']=ojb["productivity_17"]
            worksheet['G73']=ojb["productivity_18"]
            worksheet['G74']=ojb["productivity_19"]
            worksheet['G75']=ojb["productivity_20"]
            worksheet['G77']=ojb["activities_1"]
            worksheet['G78']=ojb["activities_2"]
            worksheet['G79']=ojb["activities_3"]
            worksheet['G81']=ojb["personal_1"]
            worksheet['G82']=ojb["personal_2"]
            worksheet['G83']=ojb["personal_3"]
            worksheet['G84']=ojb["personal_4"]
            worksheet['G85']=ojb["personal_5"]
            worksheet['G86']=ojb["personal_6"]
            worksheet['G87']=ojb["personal_7"]
            worksheet['G88']=ojb["personal_8"]
            worksheet['G89']=ojb["personal_9"]
            worksheet['G90']=ojb["personal_10"]
            worksheet['G91']=ojb["personal_11"]
            worksheet['G92']=ojb["personal_12"]
            worksheet['G93']=ojb["personal_13"]
            worksheet['G94']=ojb["personal_14"]
            worksheet['G95']=ojb["personal_15"]
            worksheet['G96']=ojb["personal_16"]
            worksheet['G97']=ojb["personal_17"]
            worksheet['G98']=ojb["personal_18"]
            worksheet['G99']=ojb["personal_19"]
            worksheet['G100']=ojb["personal_20"]
            worksheet['G101']=ojb["personal_21"]
            worksheet['G102']=ojb["personal_22"]
            worksheet['G103']=ojb["personal_23"]

            # حفظ ملف Excel المعدل في ذاكرة مؤقتة
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)

            # إرجاع ملف Excel كاستجابة للطلب
            response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            timeNow = datetime.datetime.now().strftime('%Y%m%d%H:%M:%S')
            name_file="report-{}-{}.xlsx".format(get_data.year,timeNow)
            print(name_file)
            response['Content-Disposition'] = f'attachment; filename="{name_file}"'
            return response



import openpyxl
from io import BytesIO
from django.shortcuts import HttpResponse
from django.http import HttpResponseServerError
import openpyxl
from io import BytesIO
from django.shortcuts import HttpResponse
from django.http import HttpResponseServerError

import openpyxl
from io import BytesIO
from django.shortcuts import HttpResponse
from django.http import HttpResponseServerError

def dssdownload_excel(request, id_Supervisory):
    """
    Downloads an Excel report based on the provided SupervisoryVisits ID.

    Args:
        request: The Django HTTP request object.
        id_Supervisory (int): The ID of the SupervisoryVisits record.

    Returns:
        HttpResponse: An HTTP response with the generated Excel file attached.
    """

    # Check if id_Supervisory is provided
    if not id_Supervisory:
        return HttpResponseBadRequest("Missing SupervisoryVisits ID")

    try:
        # Get SupervisoryVisits data
        get_data = SupervisoryVisits.objects.filter(id=id_Supervisory).first()

        # Check if data exists
        if not get_data:
            return HttpResponseNotFound("SupervisoryVisits not found")

        # Load template workbook
        template_file = 'static/template.xlsx'
        workbook = openpyxl.load_workbook(template_file, data_only=False)
        worksheet = workbook.active

        # Fill worksheet data
        worksheet.merge_cells('A4:H4')
        worksheet.cell(row=4, column=1, value=f"الفصل التدريبي {get_data.year}")

        worksheet.merge_cells('A5:H5')
        worksheet.cell(row=5, column=1, value='مجلس التدريب التقني والمهني اليمن')

        # Iterate through evaluation data using dictionary comprehension
        evaluation_data = get_data.instructorevaluation_set.values(
           "planning_1",
                "planning_2",
                "competencies_1",
                "competencies_2",
                "competencies_3",
                "competencies_4",
                "competencies_5",
                "competencies_6",
                "competencies_7",
                "competencies_8",
                "competencies_9",
                "competencies_10",
                "competencies_11",
                "competencies_12",
                "competencies_13",
                "competencies_14",
                "competencies_20",
                "competencies_21",
                "competencies_22",
                "competencies_23",
                "competencies_24",
                "development_1",
                "development_2",
                "development_3",
                "development_4",
                "development_5",
                "development_6",
                "productivity_1",
                "productivity_2",
                "productivity_3",
                "productivity_4",
                "productivity_5",
                "productivity_6",
                "productivity_7",
                "productivity_8",
                "productivity_9",
                "productivity_10",
                "productivity_11",
                "productivity_12",
                "productivity_13",
                "productivity_14",
                "productivity_15",
                "productivity_16",
                "productivity_17",
                "productivity_18",
                "productivity_19",
                "productivity_20",
                "activities_1",
                "activities_2",
                "activities_3",
                "personal_1",
                "personal_2",
                "personal_3",
                "personal_4",
                "personal_5",
                "personal_6",
                "personal_7",
                "personal_8",
                "personal_9",
                "personal_10",
                "personal_11",
                "personal_12",
                "personal_13",
                "personal_14",
                "personal_15",
                "personal_16",
                "personal_17",
                "personal_18",
                "personal_19",
                "personal_20",
                "personal_21",
                "personal_22",
                "personal_23"
        ).first()

        # Fill the worksheet with evaluation data
        row = 18
        for field, value in evaluation_data.items():
            worksheet.cell(row=row, column=7, value=value)
            row += 1

        # Save Excel file to buffer
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        # Return response with Excel attachment
        response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="output.xlsx"'
        return response

    except Exception as e:
        # Handle any errors during processing
        print(f"Error generating Excel: {e}")
        return HttpResponseServerError("Error generating Excel report")